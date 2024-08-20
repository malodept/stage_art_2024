import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Charger le fichier Excel existant
input_file = r"C:/Users/m-de-pastor/Desktop/Un peut tout/fichiers_concatenes_colonnes_gerees.xlsx"
df = pd.read_excel(input_file)

# Identifier les pays uniques
pays_uniques = df['Pays'].unique()

# Créer une nouvelle feuille Excel avec les pays en colonnes et les métadonnées en lignes
metadonnees = df.columns.drop('Pays')
result_df = pd.DataFrame(index=metadonnees, columns=pays_uniques)

# Parcourir les pays et les métadonnées pour remplir le DataFrame
for pays in pays_uniques:
    subset = df[df['Pays'] == pays]
    for metadonnee in metadonnees:
        if subset[metadonnee].notna().any():  # Si au moins une case est remplie
            result_df.at[metadonnee, pays] = 'x'  # Utilise 'x' temporairement

# Créer un nouveau fichier Excel et appliquer la mise en forme
output_file = r"C:/Users/m-de-pastor/Desktop/Un peut tout/resume_colonnes_pays.xlsx"
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    result_df.to_excel(writer, sheet_name='Résultat')
    workbook = writer.book
    worksheet = writer.sheets['Résultat']

    fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    
    for row in range(2, len(result_df) + 2):  # Commence à la 2ème ligne (après les en-têtes)
        for col in range(2, len(pays_uniques) + 2):  # Commence à la 2ème colonne (après les en-têtes)
            cell_value = worksheet.cell(row=row, column=col).value
            if cell_value == 'x':  # Remplir en bleu si la case est marquée
                worksheet.cell(row=row, column=col).fill = fill
                worksheet.cell(row=row, column=col).value = ''  # Supprimer le 'x'

print(f"Le fichier résumé a été créé avec succès : {output_file}")
